#attempt at writing a program to check for duplicates among the EC inventories
from openpyxl import load_workbook

wsx = load_workbook("ecinventoryexcelfiles/ECWSRI.xlsx")
ws = wsx.active

hlx = load_workbook("ecinventoryexcelfiles/ECHLRI.xlsx")
hl = hlx.active

senx = load_workbook("ecinventoryexcelfiles/ECSENRI.xlsx")
sen = senx.active

cpx = load_workbook("ecinventoryexcelfiles/ECC4PRI.xlsx")
cp = cpx.active

TITLE = "A"
start = 2


def list_of_titles(workbook, tlist):
    for row in workbook.range("%s%s:%s%s" % (TITLE, start, TITLE, workbook.get_highest_row())):
        for cell in row:
            tlist.append(cell.value)

wst = []
hlt = []
sent = []
cpt = []            
            
list_of_titles(ws, wst)
list_of_titles(hl, hlt)
list_of_titles(sen, sent)
list_of_titles(cp, cpt)

duplicates = []
inventories = [cpt, hlt, sent, wst]


def dupcheck(list1, list2, dup_list):
    for value in list1:
        if value in list2 and value not in dup_list:
            dup_list.append(value)

#still need to find a more efficient way to go through lists, or make a more efficient method to do this
#dupcheck(sent, wst, duplicates)
#dupcheck(sent, hlt, duplicates)
#dupcheck(sent, cpt, duplicates)
#dupcheck(wst, hlt, duplicates)
#dupcheck(wst, cpt, duplicates)
#dupcheck(hlt, cpt, duplicates)
#method below pretty much just does ^this


def multiple_dupcheck(mult_inventor, dl):
    # remaining inventories
    """

    :param mult_inventor: multiple inventories
    :param dl: duplicate list
    """
    remain_inventor = []

    for value in mult_inventor:
        remain_inventor.append(value)
        #have to do it ^this way so that list l stays intact even after the method is over
    while len(remain_inventor) > 1:
        inventory = remain_inventor[0]
        remain_inventor.remove(remain_inventor[0])
        # inventory is a single inventory in remaining inventories
        for other_inventory in remain_inventor:
            dupcheck(inventory, other_inventory, dl)

#found a way to indicate which lists the duplicates show up in, though it's a bit convoluted
#prints a list of all duplicate resources along with a four digit id consisting of 1's and 0's
#each digit corresponds with an inventory: [C4P, High Level, SEN, Workshop]
#a 1 indicates that the resource is present in the inventory, a 0 indicates that it is not


def present_in_lists(resources, lists):
    ids = dict()
    # value is a whole inventory
    for value in resources:
        itemid = str()
        for l in lists:
            if value in l:
                itemid += "1"
            else:
                itemid += "0"
        #ids.append(value + " : " + itemid)
        ids[str(value)] = itemid
    return ids

#plan on trying to figure out a way to integrate present_in_lists and the duplicate checker into one method

# inventories is a list of lists
multiple_dupcheck(inventories, duplicates)
resource_ids = present_in_lists(duplicates, inventories)
print('                          C4P  High Level  SEN  Workshop')     
for k in resource_ids:
    print(k + ': ' + resource_ids[k])
print(len(duplicates))