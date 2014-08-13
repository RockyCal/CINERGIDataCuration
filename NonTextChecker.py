import urllib.request
import time
from openpyxl import load_workbook
from bs4 import BeautifulSoup

#excel file versions of inventories
senx = load_workbook("detailsSEN.xlsx")
sen = senx.active
senlinks = []

cpx = load_workbook("detailsC4P.xlsx")
cp = cpx.active
cplinks = []

hlx = load_workbook("detailsHL.xlsx")
hl = hlx.active
hllinks = []

wsx = load_workbook("detailsWS.xlsx")
ws = wsx.active
wslinks = []

workbooks = [cp, hl, sen, ws]
urllists = [cplinks, hllinks, senlinks, wslinks]

URL = "A"
start = 1

def list_of_links(workbook, llist):
    for row in workbook.range("%s%s:%s%s" % (URL, start, URL, workbook.get_highest_row())):
        for cell in row:
            llist.append(cell.value)

#returns list of non-text symbols
def multiplefilechecker(llist):
    newlist = []
    for link in llist:
        soup = BeautifulSoup(urllib.request.urlopen(link))
        add = filechecker(soup)
        if len(add) > 1:
            newlist.append(add)
    print(len(newlist), " resources containing at least one non-text symbol")
    return newlist

def filechecker(soup):
    fields = soup.find_all("td", limit = 4) #returns text of title, brief description, url, and abstract or purpose, respectively
    newstring = str(soup.find("td"))
    for item in fields:
        string = str(item)
        if fields.index(item) != 2:
            i = -5
            for char in string:
                i += 1
                if ord(char) > 126:
                    if fields.index(item) == 0:
                        newstring += "-title"
                    elif fields.index(item) == 1:
                        newstring += "-description"
                    elif fields.index(item) == 3:
                        newstring += "-abstract"
                    newstring += str(i)                
    if len(newstring) == len(str(soup.find("td"))):
        newstring = ""
    newstring = newstring.replace("<td>", "")
    newstring = newstring.replace("</td>", "")
    return newstring

start_time = time.time()

count = 0
while count < 4:
    list_of_links(workbooks[count], urllists[count])
    count += 1

print("Non-text symbols in C4P:")
for item in multiplefilechecker(cplinks):
    print(item)
print()
print("Non-text symbols in HLI:")
for item in multiplefilechecker(hllinks):
    print(item)
print()
print("Non-text symbols in SEN:")
for item in multiplefilechecker(senlinks):
    print(item)
print()
print("Non-text symbols in WSI:")
for item in multiplefilechecker(wslinks):
    print(item)
print()

elapsed_time = time.time() - start_time
print(elapsed_time)