from openpyxl import load_workbook, cell
wb = load_workbook("EarthCube Resources.xlsx")
ws = wb.active

RESOURCE = "M"
rTitles = []
start = 2
end = ws.get_highest_row()

for row in ws.range("%s%s:%s%s"%(RESOURCE, start, RESOURCE, end)):
	for cell in row:
		rTitles.append(cell.value)

wbo = load_workbook("ECWSResourcesOnline.xlsx") 
wso = wbo.active

TITLE = "A"
# comment for reasons 
titles = []
starto = 2
endo = wso.get_highest_row()

for row in wso.range("%s%s:%s%s"%(TITLE, starto, TITLE, endo)):
	for cell in row:
		titles.append(cell.value)
		
# goes through values in list1, values in list1 not found in list2 added to list3
def compare(list1, list2, list3):
	matches = 0
	for value in list1:
		if value not in list2:
			list3.append(value)
            
notUpdated = []

compare(rTitles, titles, notUpdated)
print(notUpdated)

